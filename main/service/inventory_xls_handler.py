import openpyxl
import datetime
import time

from django.http import HttpResponse
from ..models import Inventory, InventoryOrderRow

# 発注書出力
def create_ordersheet(inventory_id):
    inventory = Inventory.objects.get(pk=inventory_id)
    company = inventory.company
    seller = inventory.seller

    # Excelのテンプレートファイルの読み込み
    wb = openpyxl.load_workbook('/django/main/static/tpl/OrderSheet.xlsx')

    sheet = wb['注文書']
    sheet['J1'] = inventory_id
    sheet['J2'] = datetime.date.today().strftime("%Y-%m-%d")

    sheet['K5'] = company.name
    sheet['K6'] = f"〒 {str(company.zip)[0:3]}-{str(company.zip)[3:]}"
    sheet['K7'] = company.address
    sheet['K9'] = f"電話：{company.tel}　FAX：{company.fax}"

    if seller:
        sheet['B5'] = f"{seller.name} 御中"
        sheet['B7'] = f"〒 {str(seller.zip)[0:3]}-{str(seller.zip)[3:]}"
        sheet['B8'] = seller.address
        sheet['B10'] = sheet['K9'] = f"電話：{seller.tel}　FAX：{seller.fax}"

    sheet['B23'] = inventory.name
    sheet['B24'] = inventory.serial_no
    sheet['H23'] = inventory.order_price

    # Excelを返すためにcontent_typeに「application/vnd.ms-excel」をセットします。
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % str(inventory.name) + '_' + str(time.time()) + '.xlsx'
    
    # データの書き込みを行なったExcelファイルを保存する
    wb.save(response)

    return response

# 請求書出力
def create_jpinvoice(inventory_id):

    # Excelのテンプレートファイルの読み込み
    wb = openpyxl.load_workbook('/django/main/static/tpl/JpInvoice.xlsx')

    sheet = wb['見積書']
    sheet_inv = wb['請求書']

    sheet['K1'] = int(time.time())
    sheet['K2'] = datetime.date.today().strftime("%Y-%m-%d")

    inventory = Inventory.objects.get(pk=inventory_id)
    company = inventory.company
    buyer = inventory.buyer

    sheet['L6'] = company.name
    sheet['L7'] = f"〒 {str(company.zip)[0:3]}-{str(company.zip)[3:]}"
    sheet['L8'] = company.address
    sheet['L9'] = f"電話：{company.tel}"
    sheet['L10'] = f"FAX：{company.fax}"
    sheet_inv['L6'] = f"登録番号 {company.registration_number}"

    if buyer:
        sheet['C2'] = f"〒 {str(buyer.zip)[0:3]}-{str(buyer.zip)[3:]}"
        sheet['C3'] = buyer.address
        sheet['C4'] = f"{buyer.name}　御中"
        sheet['C5'] = f"{buyer.pic}　様"

    # 在庫販売明細
    inventory_order_rows =  InventoryOrderRow.objects.filter(inventory_id=inventory_id)

    row=18
    for inventory_order_row in inventory_order_rows:
        # 出力対象のみ
        if inventory_order_row.is_out == False:
            continue

        sheet['B'+str(row)] = inventory_order_row.name
        sheet['G'+str(row)] = inventory_order_row.count
        sheet['I'+str(row)] = inventory_order_row.price
        sheet['J'+str(row)] = inventory_order_row.count * inventory_order_row.price
        sheet['K'+str(row)] = inventory_order_row.memo
        row = row + 2

    # Excelを返すためにcontent_typeに「application/vnd.ms-excel」をセットします。
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=invoice_' + str(time.time()) + '.xlsx'

    # データの書き込みを行なったExcelファイルを保存する
    wb.save(response)

    return response

# Proforma Invoice出力
def create_proforma_invoice(inventory_id):
    # Excelのテンプレートファイルの読み込み
    wb = openpyxl.load_workbook('/django/main/static/tpl/ProformaInvoice.xlsx')

    sheet = wb['Proforma']

    sheet['D6'] = int(time.time())
    sheet['I5'] = datetime.date.today().strftime("%Y-%m-%d")

    start_row = 18
    inventory = Inventory.objects.get(pk=inventory_id)
    company = inventory.company
    buyer = inventory.buyer
    
    sheet['B1'] = company.name_en
    sheet['B2'] = company.address_en
    sheet['B3'] = f"TEL: {company.tel_en}   FAX: {company.fax_en}"

    if buyer:
        sheet['D7'] = f"{buyer.name}"
        sheet['D8'] = buyer.address
        sheet['D10'] = f"TEL: {buyer.tel}"
        sheet['F10'] = f"FAX: {buyer.fax}"

    sheet[f"B{str(start_row)}"] = f"MODEL:{inventory.name}"
    sheet[f"B{str(start_row + 1)}"] = f"S/N:{inventory.serial_no}"
    sheet[f"F{str(start_row)}"] = 1
    sheet[f"G{str(start_row)}"] = "UNIT"
    sheet[f"H{str(start_row)}"] = inventory.sell_price

    # Excelを返すためにcontent_typeに「application/vnd.ms-excel」をセットします。
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=profprma_invoice_' + str(time.time()) + '.xlsx'

    # データの書き込みを行なったExcelファイルを保存する
    wb.save(response)

    return response
